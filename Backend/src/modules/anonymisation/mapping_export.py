from __future__ import annotations
from src.utils.logger import get_logger
logger = get_logger(__name__)

import os
from datetime import datetime
from typing import List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .entity_fields import entity_type_category_map

#  Colour palette 
_HEADER_BG = "1F3864"
_HEADER_FG = "FFFFFF"
_ROW_ODD   = "EEF2FF"
_ROW_EVEN  = "FFFFFF"
_ACCENT    = "2F75B6"

_ENTITY_COLORS: dict[str, str] = {
    "AADHAAR":           "FFF2CC",
    "PAN":               "FCE4D6",
    "PHONE":             "E2EFDA",
    "EMAIL":             "DDEBF7",
    "DATE":              "F4CCCC",
    "AGE":               "D9EAD3",
    "IP_ADDRESS":        "EAD1DC",
    "LOCATION":          "CFE2F3",
    "DIAGNOSIS":         "FFD966",
    "MEDICAL_CONDITION": "FFD966",
    "PASSPORT":          "D9D2E9",
    "BANK_ACCOUNT":      "F9CB9C",
    "PATIENT_ID":        "B6D7A8",
    "PERSON":            "EA9999",
    "ORGANISATION":      "B4A7D6",
    "MISC":              "D9D9D9",
}
_DEFAULT_COLOR = "F3F3F3"

_SOURCE_COLORS: dict[str, str] = {
    "regex":  "E2EFDA",
    "ner":    "DDEBF7",
    "hybrid": "FFF2CC",
}

_ENTITY_DESCRIPTIONS: dict[str, str] = {
    "AADHAAR":           "Indian national ID (12-digit)",
    "PAN":               "Indian tax ID (AAAAA9999A)",
    "PHONE":             "Mobile / telephone number",
    "EMAIL":             "Email address",
    "DATE":              "Date of birth / clinical date",
    "AGE":               "Patient or relative age",
    "BANK_ACCOUNT":      "Bank account / numeric ID",
    "PATIENT_ID":        "MRN / Patient ID reference",
    "PERSON":            "Personal name",
    "ORGANISATION":      "Organisation name",
    "LOCATION":          "Address / geographic location",
    "DIAGNOSIS":         "Medical diagnosis / condition",
    "MEDICAL_CONDITION": "Medical condition",
    "PASSPORT":          "Passport number",
    "IP_ADDRESS":        "IP address",
    "MISC":              "Miscellaneous PII",
}

_CATEGORY_DESC: dict[str, str] = {
    "PII":     "Personally Identifiable Information (PII)",
    "PHI":     "Protected Health Information (PHI)",
    "PII+PHI": "Mixed identifier (PII + PHI context)",
}


#  Style helpers 

def _border(color: str = "BFBFBF", style: str = "thin") -> Border:
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)


def _header_cell(ws, row: int, col: int, value: str,
                 span: int = 1, height: int = 30):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = Font(name="Arial", bold=True, color=_HEADER_FG, size=11)
    c.fill      = PatternFill("solid", start_color=_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border    = _border(_ACCENT, "medium")
    if span > 1:
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row,   end_column=col + span - 1)
    ws.row_dimensions[row].height = height
    return c


#  Sheet builders 

def _build_mapping_sheet(ws, entries, entity_lookup: dict):
    """Sheet 1 — full token ↔ original mapping table."""
    cols   = ["#", "Pseudo Token", "Original Value", "Entity Type"]
    widths = [5,    26,             34,               20]

    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        _header_cell(ws, 1, ci, h)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    for ri, entry in enumerate(entries, 2):
        bg      = _ROW_ODD if ri % 2 == 0 else _ROW_EVEN
        etype   = entry.entity_type

        row_vals = [ri - 1, entry.token, entry.original_value, etype]

        for ci, val in enumerate(row_vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font   = Font(name="Arial", size=10)
            c.border = _border()

            if ci == 1:                          # row number
                c.fill      = PatternFill("solid", start_color=_HEADER_BG)
                c.font      = Font(name="Arial", size=10, bold=True, color=_HEADER_FG)
                c.alignment = Alignment(horizontal="center", vertical="center")

            elif ci == 2:                        # pseudo token
                c.fill      = PatternFill("solid", start_color=bg)
                c.font      = Font(name="Courier New", size=9,
                                   bold=True, color="1F3864")
                c.alignment = Alignment(vertical="center")

            elif ci == 3:                        # original value
                c.fill      = PatternFill("solid", start_color=bg)
                c.alignment = Alignment(wrap_text=True, vertical="center")

            elif ci == 4:                        # entity type — colour coded
                type_bg = _ENTITY_COLORS.get(etype, _DEFAULT_COLOR)
                c.fill      = PatternFill("solid", start_color=type_bg)
                c.font      = Font(name="Arial", size=10, bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[ri].height = 18

    # Update auto-filter to include all data rows
    ws.auto_filter.ref = (
        f"A1:{get_column_letter(len(cols))}{len(entries) + 1}"
    )


def _build_summary_sheet(ws, entries, meta: dict):
    """Sheet 2 — stats dashboard."""
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["D"].width = 26
    ws.column_dimensions["E"].width = 16

    # Title
    t = ws.cell(row=1, column=1,
                value="PHI / PII Pseudo-Anonymisation Report")
    t.font = Font(name="Arial", bold=True, size=16, color=_ACCENT)
    ws.merge_cells("A1:E1")
    ws.row_dimensions[1].height = 38

    #  Left block: overview 
    _header_cell(ws, 3, 1, "Overview", span=2, height=26)

    entity_summary: dict[str, int] = {}
    source_summary: dict[str, int] = {}
    for e in entries:
        entity_summary[e.entity_type] = entity_summary.get(e.entity_type, 0) + 1
        src = getattr(e, "source", "unknown")
        source_summary[src] = source_summary.get(src, 0) + 1

    overview_rows = [
        ("Total Unique Tokens",     len(entries)),
        ("Unique Entity Types",      len(entity_summary)),
        ("Input Text Length",
         f"{meta.get('text_length', '—'):,} chars"
         if isinstance(meta.get('text_length'), int)
         else meta.get("text_length", "—")),
        ("Anonymisation Mode",       meta.get("mode", "pseudo")),
        ("Salt / HMAC",              "Yes" if meta.get("salt") else "No (random)"),
        ("Generated",                datetime.now().strftime("%d %b %Y  %H:%M")),
    ]

    for i, (label, value) in enumerate(overview_rows, 4):
        bg = _ROW_ODD if i % 2 == 0 else _ROW_EVEN
        lc = ws.cell(row=i, column=1, value=label)
        vc = ws.cell(row=i, column=2, value=value)
        lc.font      = Font(name="Arial", bold=True, size=10)
        vc.font      = Font(name="Arial", size=10)
        lc.fill      = vc.fill = PatternFill("solid", start_color=bg)
        lc.border    = vc.border = _border()
        lc.alignment = Alignment(vertical="center")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[i].height = 20

    #  Right block: entity breakdown 
    _header_cell(ws, 3, 4, "Tokens by Entity Type", span=2, height=26)

    for i, (etype, count) in enumerate(
            sorted(entity_summary.items(), key=lambda x: -x[1]), 4):
        bg = _ENTITY_COLORS.get(etype, _DEFAULT_COLOR)
        lc = ws.cell(row=i, column=4, value=etype)
        vc = ws.cell(row=i, column=5, value=count)
        lc.font   = Font(name="Arial", bold=True, size=10)
        vc.font   = Font(name="Arial", size=10)
        lc.fill   = vc.fill = PatternFill("solid", start_color=bg)
        lc.border = vc.border = _border()
        lc.alignment = Alignment(vertical="center")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[i].height = 20

    #  Right block: source breakdown 
    src_start = 4 + len(entity_summary) + 2
    _header_cell(ws, src_start, 4, "Detections by Source", span=2, height=26)

    for i, (src, count) in enumerate(
            sorted(source_summary.items(), key=lambda x: -x[1]),
            src_start + 1):
        bg = _SOURCE_COLORS.get(src.lower(), _DEFAULT_COLOR)
        lc = ws.cell(row=i, column=4, value=src.upper())
        vc = ws.cell(row=i, column=5, value=count)
        lc.font      = Font(name="Arial", bold=True, italic=True, size=10)
        vc.font      = Font(name="Arial", size=10)
        lc.fill      = vc.fill = PatternFill("solid", start_color=bg)
        lc.border    = vc.border = _border()
        lc.alignment = Alignment(vertical="center")
        vc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[i].height = 20


def _build_legend_sheet(ws, entries):
    """Sheet 3 — entity colour legend."""
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 38

    for ci, h in enumerate(["Entity Type", "Count", "Colour", "Description"], 1):
        _header_cell(ws, 1, ci, h, height=28)

    entity_summary: dict[str, int] = {}
    for e in entries:
        entity_summary[e.entity_type] = entity_summary.get(e.entity_type, 0) + 1

    for ri, (etype, count) in enumerate(
            sorted(entity_summary.items(), key=lambda x: -x[1]), 2):
        bg   = _ENTITY_COLORS.get(etype, _DEFAULT_COLOR)
        desc = _ENTITY_DESCRIPTIONS.get(etype, "")
        if not desc:
            cat = entity_type_category_map().get(etype)
            if cat:
                desc = _CATEGORY_DESC.get(cat, cat)
        vals = [etype, count, "", desc]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font   = Font(name="Arial", size=10, bold=(ci == 1))
            c.fill   = PatternFill("solid", start_color=bg)
            c.border = _border()
            if ci == 2:
                c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[ri].height = 20


def _build_reidentify_sheet(ws, entries):
    """Sheet 4 — re-identification lookup (audit use only)."""
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 20

    for ci, h in enumerate(["Pseudo Token", "Original Value", "Entity Type"], 1):
        _header_cell(ws, 1, ci, h, height=28)

    ws.freeze_panes = "A2"

    # Add a prominent warning
    warn = ws.cell(row=1, column=5,
                   value="⚠  RESTRICTED — FOR AUTHORISED AUDIT USE ONLY")
    warn.font = Font(name="Arial", bold=True, size=11, color="FF0000")

    for ri, entry in enumerate(entries, 2):
        bg    = _ROW_ODD if ri % 2 == 0 else _ROW_EVEN
        etype = entry.entity_type
        vals  = [entry.token, entry.original_value, etype]

        for ci, val in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font   = Font(name="Arial", size=10)
            c.border = _border()
            c.fill   = PatternFill("solid", start_color=bg)

            if ci == 1:
                c.font = Font(name="Courier New", size=9,
                              bold=True, color="1F3864")
            elif ci == 3:
                type_bg = _ENTITY_COLORS.get(etype, _DEFAULT_COLOR)
                c.fill  = PatternFill("solid", start_color=type_bg)
                c.font  = Font(name="Arial", size=10, bold=True)
                c.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[ri].height = 18

    ws.auto_filter.ref = f"A1:C{len(entries) + 1}"

    # Tab colour to flag sensitivity
    ws.sheet_properties.tabColor = "FF0000"


#  Public API 

def export_mapping_to_excel(
    mapping_entries,
    output_path: str,
    original_text: str = "",
    meta: Optional[dict] = None,
    entity_lookup: Optional[dict] = None,
) -> str:
    """
    Build a 4-sheet Excel workbook from mapping_entries and save to output_path.

    Parameters
    ----------
    mapping_entries : list[MappingEntry]
        The entries from MappingTableResponse.entries.
    output_path : str
        Full path for the .xlsx file (e.g. "outputs/mapping_20260406.xlsx").
    original_text : str
        The original document text (used for stats).
    meta : dict, optional
        Extra metadata: {"mode": "pseudo", "salt": True/False}.
    entity_lookup : dict, optional
        Map of original_value → DetectedEntity for score/source enrichment.

    Returns
    -------
    str  — the output_path that was written.
    """
    if meta is None:
        meta = {}
    if entity_lookup is None:
        entity_lookup = {}

    meta.setdefault("text_length", len(original_text))

    # Enrich entries with score/source from entity_lookup
    for entry in mapping_entries:
        det = entity_lookup.get(entry.original_value)
        if det and not hasattr(entry, "source"):
            entry.source = det.source
            entry.score  = round(det.score, 3)
        elif not hasattr(entry, "source"):
            entry.source = "—"
            entry.score  = "—"

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Mapping Table"
    _build_mapping_sheet(ws1, mapping_entries, entity_lookup)

    ws2 = wb.create_sheet("Summary")
    _build_summary_sheet(ws2, mapping_entries, meta)

    ws3 = wb.create_sheet("Entity Legend")
    _build_legend_sheet(ws3, mapping_entries)

    ws4 = wb.create_sheet("Re-identification (Audit)")
    _build_reidentify_sheet(ws4, mapping_entries)

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    return output_path